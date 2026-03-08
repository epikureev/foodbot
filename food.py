import os

TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")

import asyncio
import datetime
import json
import sqlite3
import time
import io

import pandas as pd
import schedule
from PIL import Image

from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command
from aiogram.types import Message, BufferedInputFile

from google import genai
from google.genai import types

from apscheduler.schedulers.asyncio import AsyncIOScheduler


# =====================
# CONFIG
# =====================


ADMIN_ID = 215444830

TEXT_MODEL = "gemini-2.0-flash"
IMAGE_MODEL = "gemini-2.5-flash"

DAILY_LIMIT = 2200


# =====================
# TELEGRAM
# =====================

bot = Bot(token=TELEGRAM_TOKEN)
dp = Dispatcher()


# =====================
# GEMINI
# =====================

from google import genai

client = genai.Client(
    api_key="dummy",
    http_options={
        "base_url": "https://bold-fire-af65.dkhabiev.workers.dev",
        "headers": {"x-bot-secret": "foodbot926346"},
    },
)

# =====================
# DATABASE
# =====================

conn = sqlite3.connect("food.db", check_same_thread=False)
cursor = conn.cursor()

cursor.execute(
    """
CREATE TABLE IF NOT EXISTS food(
id INTEGER PRIMARY KEY AUTOINCREMENT,
user_id INTEGER,
date TEXT,
food TEXT,
grams REAL,
kcal REAL,
protein REAL,
fat REAL,
carbs REAL
)
"""
)

conn.commit()


# =====================
# ERROR
# =====================


def fatal_error(e):

    print("FATAL ERROR:", e)

    os._exit(1)


# =====================
# JSON PARSER
# =====================


def parse_json(text):

    if not text:
        return None

    try:

        start = text.find("{")
        end = text.rfind("}") + 1

        if start == -1 or end == -1:
            return None

        data = json.loads(text[start:end])

        return {
            "food": data.get("food", "unknown"),
            "grams": float(data.get("grams", 0)),
            "kcal": float(data.get("kcal", 0)),
            "protein": float(data.get("protein", 0)),
            "fat": float(data.get("fat", 0)),
            "carbs": float(data.get("carbs", 0)),
        }

    except:
        return None


# =====================
# GEMINI TEXT
# =====================


def gemini_parse(text):

    prompt = """
You are a nutrition calculator.

Return ONLY JSON.

{
"food":"name",
"grams":number,
"kcal":number,
"protein":number,
"fat":number,
"carbs":number
}
"""

    models = ["gemini-2.5-flash", "gemini-2.5-flash-lite"]

    for model in models:

        for attempt in range(3):

            try:

                response = client.models.generate_content(
                    model=model, contents=f"{prompt}\nFood: {text}"
                )

                if response.text:

                    print("MODEL USED:", model)

                    return response.text

            except Exception as e:

                print("MODEL FAILED:", model, e)

                time.sleep(2)

    fatal_error("Gemini unavailable")


# =====================
# GEMINI IMAGE
# =====================


def gemini_parse_image(img):

    prompt = """
You are a professional nutrition expert and food analyst.

Your task is to analyze a food photo and estimate nutrition values.

IMPORTANT RULES:

1. The "food" field MUST always be written in Russian.
2. Use simple food names (examples: "яблоко", "омлет", "рис", "курица", "банан").
3. Do NOT include brand names or descriptions.
4. Estimate portion size realistically using common serving sizes.
5. Always return valid JSON.
6. Never include text outside JSON.

Portion estimation guidelines:

apple → 160–200 g  
banana → 110–130 g  
egg → 55–65 g  
steak → 180–250 g  
plate of rice → 200–250 g  
chicken breast → 150–220 g  
salad bowl → 150–300 g

Nutrition estimation rules:

kcal = realistic calories for the portion  
protein = grams of protein  
fat = grams of fat  
carbs = grams of carbohydrates  

Return ONLY JSON in this format:

{
"food":"название блюда на русском",
"grams":number,
"kcal":number,
"protein":number,
"fat":number,
"carbs":number
}
"""

    try:

        buffer = io.BytesIO()
        img.save(buffer, format="JPEG")

        for attempt in range(3):

            try:

                response = client.models.generate_content(
                    model=IMAGE_MODEL,
                    contents=[
                        prompt,
                        types.Part.from_bytes(
                            data=buffer.getvalue(), mime_type="image/jpeg"
                        ),
                    ],
                )

                if response.text:
                    return response.text

            except Exception as e:

                print("IMAGE MODEL FAILED:", e)
                time.sleep(2)

        return None

    except Exception as e:

        fatal_error(e)


# =====================
# SAVE FOOD
# =====================


def save_food(user, food, grams, kcal, protein, fat, carbs):

    cursor.execute(
        """
INSERT INTO food(user_id,date,food,grams,kcal,protein,fat,carbs)
VALUES(?,?,?,?,?,?,?,?)
""",
        (
            user,
            datetime.date.today().isoformat(),
            food,
            grams,
            kcal,
            protein,
            fat,
            carbs,
        ),
    )

    conn.commit()


# =====================
# DAILY TOTAL
# =====================


def daily_total(date):

    df = pd.read_sql_query(
        "SELECT kcal,protein,fat,carbs FROM food WHERE date=?", conn, params=(date,)
    )

    if df.empty:
        return None

    return df.sum()


# =====================
# WEEK STATS
# =====================


def week_stats():

    week = (datetime.date.today() - datetime.timedelta(days=7)).isoformat()

    df = pd.read_sql_query(
        "SELECT kcal,protein,fat,carbs FROM food WHERE date>=?", conn, params=(week,)
    )

    if df.empty:
        return None

    return df.mean()


# =====================
# EXCEL
# =====================


def export_excel():

    file = "food_report.xlsx"

    # получить ВСЕ данные
    df_all = pd.read_sql_query(
        "SELECT date,food,grams,kcal,protein,fat,carbs FROM food",
        conn,
    )

    if df_all.empty:
        return None

    df_all["date"] = pd.to_datetime(df_all["date"])

    # открыть Excel
    if os.path.exists(file):
        writer = pd.ExcelWriter(
            file,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace",
        )
    else:
        writer = pd.ExcelWriter(file, engine="openpyxl")

    # --- листы по дням ---

    for date, df in df_all.groupby(df_all["date"].dt.date):

        sheet = str(date)

        df = df[["food", "grams", "kcal", "protein", "fat", "carbs"]]

        df.to_excel(writer, sheet_name=sheet, index=False)

        ws = writer.book[sheet]

        last_row = len(df) + 2

        ws[f"A{last_row}"] = "ИТОГО"

        kcal = df["kcal"].sum()
        protein = df["protein"].sum()
        fat = df["fat"].sum()
        carbs = df["carbs"].sum()

        ws[f"C{last_row}"] = kcal
        ws[f"D{last_row}"] = protein
        ws[f"E{last_row}"] = fat
        ws[f"F{last_row}"] = carbs

    # --- удалить листы старше 60 дней ---

    limit = datetime.date.today() - datetime.timedelta(days=60)

    for sheet in list(writer.book.sheetnames):

        try:
            d = datetime.datetime.strptime(sheet, "%Y-%m-%d").date()

            if d < limit:
                del writer.book[sheet]

        except:
            pass

    # --- SUMMARY таблица ---

    summary_data = []

    for sheet in writer.book.sheetnames:

        try:

            datetime.datetime.strptime(sheet, "%Y-%m-%d")

            ws = writer.book[sheet]

            row = ws.max_row

            summary_data.append(
                {
                    "date": sheet,
                    "kcal": ws[f"C{row}"].value,
                    "protein": ws[f"D{row}"].value,
                    "fat": ws[f"E{row}"].value,
                    "carbs": ws[f"F{row}"].value,
                }
            )

        except:
            pass

    summary_df = pd.DataFrame(summary_data).sort_values("date")

    summary_df.to_excel(writer, sheet_name="SUMMARY", index=False)

    # --- графики ---

    from openpyxl.chart import LineChart, Reference

    ws = writer.book["SUMMARY"]

    # --- график калорий ---

    chart = LineChart()
    chart.title = "Calories trend (60 days)"
    chart.y_axis.title = "Calories"
    chart.x_axis.title = "Date"

    data = Reference(ws, min_col=2, min_row=1, max_row=len(summary_df) + 1)
    chart.add_data(data, titles_from_data=True)

    cats = Reference(ws, min_col=1, min_row=2, max_row=len(summary_df) + 1)
    chart.set_categories(cats)

    ws.add_chart(chart, "H2")

    # --- добавить колонку лимита ---

    ws["F1"] = "limit"

    for i in range(2, len(summary_df) + 2):
        ws[f"F{i}"] = DAILY_LIMIT

    # --- график калории vs лимит ---

    chart2 = LineChart()
    chart2.title = "Calories vs Limit"
    chart2.y_axis.title = "Calories"
    chart2.x_axis.title = "Date"

    data2 = Reference(ws, min_col=2, max_col=6, min_row=1, max_row=len(summary_df) + 1)
    chart2.add_data(data2, titles_from_data=True)

    cats2 = Reference(ws, min_col=1, min_row=2, max_row=len(summary_df) + 1)
    chart2.set_categories(cats2)

    ws.add_chart(chart2, "H20")

    writer.close()

    return file


# =====================
# COMMANDS
# =====================


@dp.message(Command("start"))
async def start(message: Message):

    await message.answer(
        """
Отправьте:

текст еды
фото еды

Команды:

/undo
/stats
/week
/excel
"""
    )


# =====================
# DELETE LAST
# =====================


@dp.message(Command("undo"))
async def undo(message: Message):

    cursor.execute(
        """
DELETE FROM food
WHERE id=(
SELECT id FROM food
WHERE user_id=?
ORDER BY id DESC
LIMIT 1)
""",
        (message.from_user.id,),
    )

    conn.commit()

    await message.answer("Последняя запись удалена")


# =====================
# TODAY STATS
# =====================


@dp.message(Command("stats"))
async def stats(message: Message):

    today = datetime.date.today().isoformat()

    total = daily_total(today)

    if total is None:

        await message.answer("Сегодня записей нет")
        return

    await message.answer(
        f"""
Сегодня:

Калории {total.kcal:.0f}
Белки {total.protein:.1f}
Жиры {total.fat:.1f}
Углеводы {total.carbs:.1f}

Осталось {DAILY_LIMIT-total.kcal:.0f} kcal
"""
    )


# =====================
# WEEK
# =====================


@dp.message(Command("week"))
async def week(message: Message):

    avg = week_stats()

    if avg is None:

        await message.answer("Нет данных")
        return

    await message.answer(
        f"""
Среднее за неделю

Калории {avg.kcal:.0f}
Белки {avg.protein:.1f}
Жиры {avg.fat:.1f}
Углеводы {avg.carbs:.1f}
"""
    )


# =====================
# EXCEL
# =====================


@dp.message(Command("excel"))
async def excel(message: Message):

    df = pd.read_sql_query("SELECT * FROM food", conn)

    if df.empty:
        await message.answer("В базе пока нет данных")
        return

    buffer = io.BytesIO()

    df.to_excel(buffer, index=False)

    buffer.seek(0)

    await message.answer_document(
        BufferedInputFile(buffer.getvalue(), filename="food_report.xlsx")
    )


# =====================
# PHOTO
# =====================


@dp.message(F.photo)
async def photo(message: Message):

    photo = message.photo[-1]

    file = "food.jpg"

    await bot.download(photo, file)

    img = Image.open(file)

    img = img.resize((1024, 1024))

    result = gemini_parse_image(img)

    data = parse_json(result)

    if not data:

        await message.answer("Не смог распознать")
        return

    save_food(
        message.from_user.id,
        data["food"],
        data["grams"],
        data["kcal"],
        data["protein"],
        data["fat"],
        data["carbs"],
    )

    today = datetime.date.today().isoformat()

    total = daily_total(today)

    if total is None:
        eaten = data["kcal"]
    else:
        eaten = total.kcal

    left = DAILY_LIMIT - eaten

    await message.answer(
        f"""
🍽 {data['food']}

⚖️ Вес: {data['grams']:.0f} г
🔥 Калории: {data['kcal']:.0f} ккал

🥩 Белки: {data['protein']:.1f} г
🧈 Жиры: {data['fat']:.1f} г
🍞 Углеводы: {data['carbs']:.1f} г

📊 Осталось на сегодня: {left:.0f} ккал
"""
    )


# =====================
# TEXT
# =====================


@dp.message(F.text)
async def text(message: Message):

    wait_msg = await message.answer("⏳ Анализирую еду...")

    try:

        result = await asyncio.wait_for(
            asyncio.to_thread(gemini_parse, message.text), timeout=30
        )

    except asyncio.TimeoutError:

        await wait_msg.edit_text("⚠️ Сервер долго отвечает. Попробуйте ещё раз.")
        return

    data = parse_json(result)

    if not data:

        await wait_msg.edit_text("Ошибка распознавания")
        return

    save_food(
        message.from_user.id,
        data["food"],
        data["grams"],
        data["kcal"],
        data["protein"],
        data["fat"],
        data["carbs"],
    )

    today = datetime.date.today().isoformat()

    total = daily_total(today)

    if total is None:
        eaten = data["kcal"]
    else:
        eaten = total.kcal

    left = DAILY_LIMIT - eaten

    await wait_msg.edit_text(
        f"""
🍽 {data['food']}

⚖️ Вес: {data['grams']:.0f} г
🔥 Калории: {data['kcal']:.0f} ккал

🥩 Белки: {data['protein']:.1f} г
🧈 Жиры: {data['fat']:.1f} г
🍞 Углеводы: {data['carbs']:.1f} г

📊 Осталось на сегодня: {left:.0f} ккал
"""
    )


# =====================
# REPORT
# =====================


async def report():

    yesterday = (datetime.date.today() - datetime.timedelta(days=1)).isoformat()

    total = daily_total(yesterday)

    if total is None:
        return

    await bot.send_message(
        ADMIN_ID,
        f"""
Итого за {yesterday}

Калории {total.kcal:.0f}
Белки {total.protein:.1f}
Жиры {total.fat:.1f}
Углеводы {total.carbs:.1f}
""",
    )


# =====================
# SCHEDULER
# =====================


async def send_daily_excel():

    print("GENERATING DAILY EXCEL")

    yesterday = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime(
        "%Y-%m-%d"
    )

    conn = sqlite3.connect("food.db")

    df = pd.read_sql_query(
        "SELECT * FROM food WHERE date LIKE ?", conn, params=(f"{yesterday}%",)
    )

    conn.close()

    if df.empty:
        print("NO DATA FOR YESTERDAY")
        return

    buffer = io.BytesIO()

    df.to_excel(buffer, index=False)

    buffer.seek(0)

    await bot.send_document(
        chat_id=ADMIN_ID,
        document=BufferedInputFile(buffer.getvalue(), filename="food_report.xlsx"),
    )

    print("EXCEL SENT")


# =====================
# MAIN
# =====================


async def main():

    print("BOT STARTED")

    scheduler = AsyncIOScheduler(timezone="Europe/Istanbul")

    scheduler.add_job(send_daily_excel, trigger="cron", hour=8, minute=0)

    scheduler.start()

    print("SCHEDULER STARTED")

    await dp.start_polling(bot)


if __name__ == "__main__":

    asyncio.run(main())
